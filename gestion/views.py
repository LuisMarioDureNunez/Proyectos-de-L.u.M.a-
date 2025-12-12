# views.py - VISTAS COMPLETAS Y MEJORADAS
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth import login
from gestion import views
from django.urls import reverse_lazy, reverse
from django.views.generic import CreateView, ListView, UpdateView, DeleteView
from django.db.models import Q, Sum, Count
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.forms import modelformset_factory
import json
from django.http import HttpResponse

from django.http import JsonResponse
from django.forms import formset_factory
from decimal import Decimal
from django.http import Http404
from datetime import datetime, timedelta

from .models import Presupuesto, ItemPresupuesto, Obra
from .forms import PresupuestoForm, ItemPresupuestoForm, PresupuestoWizardForm
# from .utils.pdf_utils import PDFGenerator, ChartGenerator  # Comentado temporalmente
from .models import Presupuesto, Obra, Material
from .models import *
from .forms import *

# Importar las nuevas vistas
from .views_produccion import *
from .views_usuario import *

# =============================================
# DECORADORES DE PERMISOS
# =============================================

def es_administrador(user):
    return user.is_authenticated and user.es_administrador()

def puede_gestionar_obras(user):
    return user.is_authenticated and user.puede_gestionar_obras()

# =============================================
# VISTAS PRINCIPALES
# =============================================

def bienvenida(request):
    """Página de bienvenida con términos y condiciones"""
    return render(request, 'bienvenida.html')

def home(request):
    """Página principal del sistema - SIN REDIRECCIONES FORZADAS"""
    # NUNCA redirigir automáticamente desde home
    # Dejar que el usuario elija qué hacer
    
    context = {
        'titulo': 'Sistema de Gestión de Presupuestos - Paraguay',
        'descripcion': 'Plataforma profesional para la gestión de obras civiles y presupuestos en Paraguay',
        'caracteristicas': [
            'Gestión completa de obras civiles',
            'Presupuestos detallados y profesionales', 
            'Control de inventario de materiales',
            'Seguimiento de maquinarias y herramientas',
            'Reportes avanzados y estadísticas',
            'Sistema de roles y permisos'
        ],
        'usuario_autenticado': request.user.is_authenticated
    }
    
    # Siempre mostrar la página de inicio
    return render(request, 'home_publica.html', context)

@login_required
def dashboard(request):
    """Dashboard principal - llama directamente al dashboard super"""
    return dashboard_super(request)

@login_required
def dashboard_super(request):
    """Dashboard super con gráficos avanzados y tiempo real"""
    from decimal import Decimal
    from django.db.models import Sum, Count
    from django.utils import timezone
    from datetime import timedelta
    
    hoy = timezone.now().date()
    hace_7_dias = hoy - timedelta(days=7)
    
    stats = {
        'total_obras': Obra.objects.count(),
        'total_presupuestos': Presupuesto.objects.count(),
        'total_materiales': Material.objects.count(),
        'total_usuarios': UsuarioPersonalizado.objects.count(),
        'total_maquinarias': Maquinaria.objects.count(),
        'total_herramientas': Herramienta.objects.count(),
        'obras_planificadas': Obra.objects.filter(estado='planificada').count(),
        'obras_en_proceso': Obra.objects.filter(estado='en_proceso').count(),
        'obras_finalizadas': Obra.objects.filter(estado='finalizada').count(),
        'presupuestos_solicitados': Presupuesto.objects.filter(estado='solicitado').count(),
        'presupuestos_aceptados': Presupuesto.objects.filter(estado='aceptado').count(),
        'presupuestos_rechazados': Presupuesto.objects.filter(estado='rechazado').count(),
    }
    
    total_aceptados = Presupuesto.objects.filter(estado='aceptado').aggregate(total=Sum('total'))['total'] or Decimal('0')
    stats['total_presupuestos_aceptados'] = int(total_aceptados)
    
    obras_recientes = Obra.objects.all().order_by('-fecha_creacion')[:5]
    presupuestos_recientes = Presupuesto.objects.all().order_by('-fecha_creacion')[:5]
    
    actividades_recientes = []
    obras_nuevas = Obra.objects.filter(fecha_creacion__gte=hace_7_dias).order_by('-fecha_creacion')[:5]
    for obra in obras_nuevas:
        actividades_recientes.append({
            'tipo': 'obra',
            'accion': 'creada',
            'objeto': obra.nombre,
            'fecha': obra.fecha_creacion,
            'usuario': obra.creado_por.username if obra.creado_por else 'Sistema',
            'icono': '🏗️'
        })
    
    presupuestos_nuevos = Presupuesto.objects.filter(fecha_creacion__gte=hace_7_dias).order_by('-fecha_creacion')[:5]
    for presupuesto in presupuestos_nuevos:
        actividades_recientes.append({
            'tipo': 'presupuesto',
            'accion': 'creado',
            'objeto': f'Presupuesto {presupuesto.codigo_presupuesto}',
            'fecha': presupuesto.fecha_creacion,
            'usuario': presupuesto.cliente.username,
            'icono': '💰'
        })
    
    actividades_recientes.sort(key=lambda x: x['fecha'], reverse=True)
    
    context = {
        'stats': stats,
        'obras_recientes': obras_recientes,
        'presupuestos_recientes': presupuestos_recientes,
        'actividades_recientes': actividades_recientes[:10],
    }
    
    return render(request, 'gestion/dashboard/dashboard_super.html', context)

@login_required
def dashboard_increible(request):
    """Dashboard increíble con datos reales del proyecto"""
    from decimal import Decimal
    from django.db.models import Sum, Count
    from django.utils import timezone
    from datetime import timedelta
    
    hoy = timezone.now().date()
    hace_7_dias = hoy - timedelta(days=7)
    
    # Estadísticas principales
    stats = {
        'total_obras': Obra.objects.count(),
        'total_presupuestos': Presupuesto.objects.count(),
        'total_materiales': Material.objects.count(),
        'total_usuarios': UsuarioPersonalizado.objects.count(),
        'total_maquinarias': Maquinaria.objects.count(),
        'total_herramientas': Herramienta.objects.count(),
        'obras_planificadas': Obra.objects.filter(estado='planificada').count(),
        'obras_en_proceso': Obra.objects.filter(estado='en_proceso').count(),
        'obras_finalizadas': Obra.objects.filter(estado='finalizada').count(),
        'presupuestos_solicitados': Presupuesto.objects.filter(estado='solicitado').count(),
        'presupuestos_aceptados': Presupuesto.objects.filter(estado='aceptado').count(),
        'presupuestos_rechazados': Presupuesto.objects.filter(estado='rechazado').count(),
    }
    
    # Ingresos totales
    total_aceptados = Presupuesto.objects.filter(estado='aceptado').aggregate(total=Sum('total'))['total'] or Decimal('0')
    stats['total_presupuestos_aceptados'] = int(total_aceptados)
    
    # Obras recientes
    obras_recientes = Obra.objects.all().order_by('-fecha_creacion')[:5]
    
    # Presupuestos recientes
    presupuestos_recientes = Presupuesto.objects.all().order_by('-fecha_creacion')[:5]
    
    # Actividad reciente
    actividades_recientes = []
    obras_nuevas = Obra.objects.filter(fecha_creacion__gte=hace_7_dias).order_by('-fecha_creacion')[:5]
    for obra in obras_nuevas:
        actividades_recientes.append({
            'tipo': 'obra',
            'accion': 'creada',
            'objeto': obra.nombre,
            'fecha': obra.fecha_creacion,
            'usuario': obra.creado_por.username if obra.creado_por else 'Sistema',
            'icono': '🏗️'
        })
    
    presupuestos_nuevos = Presupuesto.objects.filter(fecha_creacion__gte=hace_7_dias).order_by('-fecha_creacion')[:5]
    for presupuesto in presupuestos_nuevos:
        actividades_recientes.append({
            'tipo': 'presupuesto',
            'accion': 'creado',
            'objeto': f'Presupuesto {presupuesto.codigo_presupuesto}',
            'fecha': presupuesto.fecha_creacion,
            'usuario': presupuesto.cliente.username,
            'icono': '💰'
        })
    
    actividades_recientes.sort(key=lambda x: x['fecha'], reverse=True)
    
    context = {
        'stats': stats,
        'obras_recientes': obras_recientes,
        'presupuestos_recientes': presupuestos_recientes,
        'actividades_recientes': actividades_recientes[:10],
    }
    
    return render(request, 'gestion/dashboard/dashboard_increible.html', context)

# =============================================
# GESTIÓN DE MATERIALES (CRUD COMPLETO)
# =============================================

@login_required
@user_passes_test(es_administrador)
def lista_materiales(request):
    """Lista todos los materiales con filtros y estadísticas"""
    materiales = Material.objects.all().order_by('-fecha_creacion')
    
    # Filtros
    query = request.GET.get('q')
    if query:
        materiales = materiales.filter(
            Q(nombre__icontains=query) | 
            Q(descripcion__icontains=query)
        )
    
    # Filtro por stock
    stock_filter = request.GET.get('stock')
    if stock_filter == 'bajo':
        materiales = materiales.filter(stock__lt=10)
    elif stock_filter == 'sin_stock':
        materiales = materiales.filter(stock=0)
    
    # Estadísticas
    total_materiales = Material.objects.count()
    stock_alto = Material.objects.filter(stock__gte=50).count()
    stock_bajo = Material.objects.filter(stock__lt=10, stock__gt=0).count()
    sin_stock = Material.objects.filter(stock=0).count()
    
    context = {
        'materiales': materiales,
        'query': query or '',
        'stock_filter': stock_filter or '',
        'total_materiales': total_materiales,
        'stock_alto': stock_alto,
        'stock_bajo': stock_bajo,
        'sin_stock': sin_stock,
    }
    return render(request, 'gestion/materiales/lista.html', context)

@login_required
@user_passes_test(es_administrador)
def nuevo_material(request):
    """Crear nuevo material"""
    if request.method == 'POST':
        form = FormularioMaterial(request.POST)
        if form.is_valid():
            material = form.save(commit=False)
            material.creado_por = request.user
            material.save()
            messages.success(request, '✅ Material creado exitosamente')
            return redirect('lista_materiales')
    else:
        form = FormularioMaterial()
    
    return render(request, 'gestion/materiales/form.html', {
        'form': form,
        'titulo': 'Nuevo Material',
        'accion': 'Crear'
    })

@login_required
@user_passes_test(es_administrador)
def editar_material(request, id):
    """Editar material existente"""
    material = get_object_or_404(Material, id=id)
    
    if request.method == 'POST':
        form = FormularioMaterial(request.POST, instance=material)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Material actualizado exitosamente')
            return redirect('lista_materiales')
    else:
        form = FormularioMaterial(instance=material)
    
    return render(request, 'gestion/materiales/form.html', {
        'form': form,
        'titulo': 'Editar Material',
        'accion': 'Actualizar',
        'material': material
    })

@login_required
@user_passes_test(es_administrador)
def eliminar_material(request, id):
    """Eliminar material"""
    material = get_object_or_404(Material, id=id)
    
    if request.method == 'POST':
        nombre_material = material.nombre
        material.delete()
        messages.success(request, f'✅ Material "{nombre_material}" eliminado exitosamente')
        return redirect('lista_materiales')
    
    return render(request, 'gestion/materiales/eliminar.html', {'material': material})

# =============================================
# GESTIÓN DE MAQUINARIAS (CRUD COMPLETO)
# =============================================

@login_required
@user_passes_test(es_administrador)
def lista_maquinarias(request):
    """Lista todas las maquinarias con estadísticas"""
    maquinarias = Maquinaria.objects.all().order_by('-fecha_creacion')
    
    query = request.GET.get('q')
    if query:
        maquinarias = maquinarias.filter(
            Q(nombre__icontains=query) | 
            Q(descripcion__icontains=query)
        )
    
    estado_filter = request.GET.get('estado')
    if estado_filter:
        maquinarias = maquinarias.filter(estado=estado_filter)
    
    # Estadísticas
    total_maquinarias = Maquinaria.objects.count()
    disponibles = Maquinaria.objects.filter(estado='disponible').count()
    en_mantenimiento = Maquinaria.objects.filter(estado='mantenimiento').count()
    alquiladas = Maquinaria.objects.filter(estado='alquilada').count()
    
    context = {
        'maquinarias': maquinarias,
        'query': query or '',
        'estado_filter': estado_filter or '',
        'total_maquinarias': total_maquinarias,
        'disponibles': disponibles,
        'en_mantenimiento': en_mantenimiento,
        'alquiladas': alquiladas,
    }
    return render(request, 'gestion/maquinarias/lista.html', context)

@login_required
@user_passes_test(es_administrador)
def nueva_maquinaria(request):
    """Crear nueva maquinaria"""
    if request.method == 'POST':
        form = FormularioMaquinaria(request.POST)
        if form.is_valid():
            maquinaria = form.save(commit=False)
            maquinaria.creado_por = request.user
            maquinaria.save()
            messages.success(request, '✅ Maquinaria creada exitosamente')
            return redirect('lista_maquinarias')
    else:
        form = FormularioMaquinaria()
    
    return render(request, 'gestion/maquinarias/form.html', {
        'form': form,
        'titulo': 'Nueva Maquinaria',
        'accion': 'Crear'
    })

@login_required
@user_passes_test(es_administrador)
def editar_maquinaria(request, id):
    """Editar maquinaria existente"""
    maquinaria = get_object_or_404(Maquinaria, id=id)
    
    if request.method == 'POST':
        form = FormularioMaquinaria(request.POST, instance=maquinaria)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Maquinaria actualizada exitosamente')
            return redirect('lista_maquinarias')
    else:
        form = FormularioMaquinaria(instance=maquinaria)
    
    return render(request, 'gestion/maquinarias/form.html', {
        'form': form,
        'titulo': 'Editar Maquinaria',
        'accion': 'Actualizar',
        'maquinaria': maquinaria
    })

@login_required
@user_passes_test(es_administrador)
def eliminar_maquinaria(request, id):
    """Eliminar maquinaria"""
    maquinaria = get_object_or_404(Maquinaria, id=id)
    
    if request.method == 'POST':
        nombre_maquinaria = maquinaria.nombre
        maquinaria.delete()
        messages.success(request, f'✅ Maquinaria "{nombre_maquinaria}" eliminada exitosamente')
        return redirect('lista_maquinarias')
    
    return render(request, 'gestion/maquinarias/eliminar.html', {'maquinaria': maquinaria})

# =============================================
# GESTIÓN DE HERRAMIENTAS (CRUD COMPLETO)
# =============================================

@login_required
@user_passes_test(es_administrador)
def lista_herramientas(request):
    """Lista todas las herramientas con estadísticas"""
    herramientas = Herramienta.objects.all().order_by('-fecha_creacion')
    
    query = request.GET.get('q')
    if query:
        herramientas = herramientas.filter(
            Q(nombre__icontains=query) | 
            Q(descripcion__icontains=query)
        )
    
    estado_filter = request.GET.get('estado')
    if estado_filter:
        herramientas = herramientas.filter(estado=estado_filter)
    
    # Estadísticas
    total_herramientas = Herramienta.objects.count()
    disponibles = Herramienta.objects.filter(estado='disponible').count()
    en_uso = Herramienta.objects.filter(estado='en_uso').count()
    en_mantenimiento = Herramienta.objects.filter(estado='mantenimiento').count()
    
    context = {
        'herramientas': herramientas,
        'query': query or '',
        'estado_filter': estado_filter or '',
        'total_herramientas': total_herramientas,
        'disponibles': disponibles,
        'en_uso': en_uso,
        'en_mantenimiento': en_mantenimiento,
    }
    return render(request, 'gestion/herramientas/lista.html', context)

@login_required
@user_passes_test(es_administrador)
def nueva_herramienta(request):
    """Crear nueva herramienta"""
    if request.method == 'POST':
        form = FormularioHerramienta(request.POST)
        if form.is_valid():
            herramienta = form.save(commit=False)
            herramienta.creado_por = request.user
            herramienta.cantidad_disponible = herramienta.cantidad_total
            herramienta.save()
            messages.success(request, '✅ Herramienta creada exitosamente')
            return redirect('lista_herramientas')
    else:
        form = FormularioHerramienta()
    
    return render(request, 'gestion/herramientas/form.html', {
        'form': form,
        'titulo': 'Nueva Herramienta',
        'accion': 'Crear'
    })

@login_required
@user_passes_test(es_administrador)
def editar_herramienta(request, id):
    """Editar herramienta existente"""
    herramienta = get_object_or_404(Herramienta, id=id)
    
    if request.method == 'POST':
        form = FormularioHerramienta(request.POST, instance=herramienta)
        if form.is_valid():
            herramienta_actualizada = form.save(commit=False)
            # Si cambia la cantidad total, actualizar la disponible
            if herramienta_actualizada.cantidad_total > herramienta.cantidad_total:
                diferencia = herramienta_actualizada.cantidad_total - herramienta.cantidad_total
                herramienta_actualizada.cantidad_disponible += diferencia
            herramienta_actualizada.save()
            messages.success(request, '✅ Herramienta actualizada exitosamente')
            return redirect('lista_herramientas')
    else:
        form = FormularioHerramienta(instance=herramienta)
    
    return render(request, 'gestion/herramientas/form.html', {
        'form': form,
        'titulo': 'Editar Herramienta',
        'accion': 'Actualizar',
        'herramienta': herramienta
    })

@login_required
@user_passes_test(es_administrador)
def eliminar_herramienta(request, id):
    """Eliminar herramienta"""
    herramienta = get_object_or_404(Herramienta, id=id)
    
    if request.method == 'POST':
        nombre_herramienta = herramienta.nombre
        herramienta.delete()
        messages.success(request, f'✅ Herramienta "{nombre_herramienta}" eliminada exitosamente')
        return redirect('lista_herramientas')
    
    return render(request, 'gestion/herramientas/eliminar.html', {'herramienta': herramienta})

# =============================================
# GESTIÓN DE OBRAS (CRUD COMPLETO)
# =============================================

@login_required
def lista_obras(request):
    """Lista de obras con filtros por rol"""
    if request.user.es_administrador() or request.user.rol == 'constructor':
        obras = Obra.objects.all().order_by('-fecha_creacion')
    else:
        # Clientes solo ven sus obras
        obras = Obra.objects.filter(cliente=request.user).order_by('-fecha_creacion')
    
    # Filtros
    estado_filter = request.GET.get('estado')
    if estado_filter:
        obras = obras.filter(estado=estado_filter)
    
    query = request.GET.get('q')
    if query:
        obras = obras.filter(
            Q(nombre__icontains=query) | 
            Q(descripcion__icontains=query) |
            Q(ubicacion__icontains=query)
        )
    
    context = {
        'obras': obras,
        'estado_filter': estado_filter or '',
        'query': query or '',
        'puede_crear': request.user.puede_gestionar_obras(),
    }
    return render(request, 'gestion/obras/lista.html', context)

@login_required
@user_passes_test(puede_gestionar_obras)
def nueva_obra(request):
    """Crear nueva obra"""
    if request.method == 'POST':
        form = FormularioObra(request.POST)
        if form.is_valid():
            obra = form.save(commit=False)
            obra.creado_por = request.user
            obra.save()
            messages.success(request, '✅ Obra creada exitosamente')
            return redirect('lista_obras')
    else:
        form = FormularioObra()
    
    return render(request, 'gestion/obras/form.html', {
        'form': form,
        'titulo': 'Nueva Obra',
        'accion': 'Crear'
    })

@login_required
@user_passes_test(puede_gestionar_obras)
def editar_obra(request, id):
    """Editar obra existente"""
    obra = get_object_or_404(Obra, id=id)
    
    if request.method == 'POST':
        form = FormularioObra(request.POST, instance=obra)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Obra actualizada exitosamente')
            return redirect('lista_obras')
    else:
        form = FormularioObra(instance=obra)
    
    return render(request, 'gestion/obras/form.html', {
        'form': form,
        'titulo': 'Editar Obra',
        'accion': 'Actualizar',
        'obra': obra
    })

@login_required
@user_passes_test(es_administrador)
def eliminar_obra(request, id):
    """Eliminar obra"""
    obra = get_object_or_404(Obra, id=id)
    
    if request.method == 'POST':
        nombre_obra = obra.nombre
        obra.delete()
        messages.success(request, f'✅ Obra "{nombre_obra}" eliminada exitosamente')
        return redirect('lista_obras')
    
    return render(request, 'gestion/obras/eliminar.html', {'obra': obra})

@login_required
def detalle_obra(request, id):
    """Detalle completo de una obra"""
    obra = get_object_or_404(Obra, id=id)
    
    # Verificar permisos
    if not request.user.es_administrador() and obra.cliente != request.user:
        messages.error(request, '❌ No tienes permisos para ver esta obra')
        return redirect('lista_obras')
    
    # Presupuestos relacionados
    presupuestos = obra.presupuestos.all().order_by('-fecha_creacion')
    
    context = {
        'obra': obra,
        'presupuestos': presupuestos,
    }
    return render(request, 'gestion/obras/detalle.html', context)

@login_required
def obras_finalizadas(request):
    """Lista de obras finalizadas"""
    if request.user.es_administrador() or request.user.rol == 'constructor':
        obras = Obra.objects.filter(estado='finalizada').order_by('-fecha_creacion')
    else:
        obras = Obra.objects.filter(cliente=request.user, estado='finalizada').order_by('-fecha_creacion')
    
    context = {
        'obras': obras,
        'titulo_especial': 'Obras Finalizadas'
    }
    return render(request, 'gestion/obras/lista.html', context)

# =============================================
# GESTIÓN DE PRESUPUESTOS (CRUD COMPLETO)
# =============================================

@login_required
def lista_presupuestos(request):
    """Lista de presupuestos con filtros por rol"""
    if request.user.es_administrador() or request.user.rol == 'constructor':
        presupuestos = Presupuesto.objects.all().order_by('-fecha_creacion')
    else:
        # Clientes solo ven sus presupuestos
        presupuestos = Presupuesto.objects.filter(cliente=request.user).order_by('-fecha_creacion')
    
    # Filtros
    estado_filter = request.GET.get('estado')
    if estado_filter:
        presupuestos = presupuestos.filter(estado=estado_filter)
    
    obra_filter = request.GET.get('obra')
    if obra_filter:
        presupuestos = presupuestos.filter(obra_id=obra_filter)
    
    context = {
        'presupuestos': presupuestos,
        'estado_filter': estado_filter or '',
        'obra_filter': obra_filter or '',
        'obras': Obra.objects.all() if request.user.es_administrador() else Obra.objects.filter(cliente=request.user),
    }
    return render(request, 'gestion/presupuestos/lista.html', context)

@login_required
def solicitar_presupuesto(request):
    """Solicitar un nuevo presupuesto (para clientes)"""
    if not request.user.rol == 'cliente':
        messages.error(request, '❌ Solo los clientes pueden solicitar presupuestos')
        return redirect('lista_presupuestos')
    
    if request.method == 'POST':
        form = FormularioPresupuesto(request.POST)
        if form.is_valid():
            presupuesto = form.save(commit=False)
            presupuesto.cliente = request.user
            presupuesto.estado = 'solicitado'
            presupuesto.save()
            messages.success(request, '✅ Presupuesto solicitado exitosamente')
            return redirect('lista_presupuestos')
    else:
        form = FormularioPresupuesto()
        # Filtrar obras del cliente
        form.fields['obra'].queryset = Obra.objects.filter(cliente=request.user, estado__in=['planificada', 'en_proceso'])
    
    return render(request, 'gestion/presupuestos/solicitar.html', {
        'form': form,
        'titulo': 'Solicitar Presupuesto'
    })

@login_required
@user_passes_test(puede_gestionar_obras)
def nuevo_presupuesto(request):
    """Crear nuevo presupuesto (para administradores/constructores)"""
    if request.method == 'POST':
        form = FormularioPresupuesto(request.POST)
        if form.is_valid():
            presupuesto = form.save(commit=False)
            presupuesto.estado = 'en_revision'
            presupuesto.save()
            messages.success(request, '✅ Presupuesto creado exitosamente')
            return redirect('lista_presupuestos')
    else:
        form = FormularioPresupuesto()
    
    return render(request, 'gestion/presupuestos/form.html', {
        'form': form,
        'titulo': 'Nuevo Presupuesto',
        'accion': 'Crear'
    })

@login_required
def detalle_presupuesto(request, id):
    """Detalle completo de un presupuesto"""
    presupuesto = get_object_or_404(Presupuesto, id=id)
    
    # Verificar permisos
    if not request.user.es_administrador() and presupuesto.cliente != request.user:
        messages.error(request, '❌ No tienes permisos para ver este presupuesto')
        return redirect('lista_presupuestos')
    
    # Items del presupuesto
    items = presupuesto.items.all()
    
    context = {
        'presupuesto': presupuesto,
        'items': items,
    }
    return render(request, 'gestion/presupuestos/detalle.html', context)

@login_required
def aceptar_presupuesto(request, id):
    """Aceptar un presupuesto (cliente)"""
    presupuesto = get_object_or_404(Presupuesto, id=id)
    
    if presupuesto.cliente != request.user:
        messages.error(request, '❌ Solo el cliente puede aceptar este presupuesto')
        return redirect('lista_presupuestos')
    
    if request.method == 'POST':
        presupuesto.estado = 'aceptado'
        presupuesto.save()
        
        # Actualizar la obra con el presupuesto aceptado
        obra = presupuesto.obra
        obra.presupuesto_asignado = presupuesto.total
        obra.estado = 'en_proceso'
        obra.save()
        
        messages.success(request, '✅ Presupuesto aceptado exitosamente')
        return redirect('lista_presupuestos')
    
    return render(request, 'gestion/presupuestos/confirmar_accion.html', {
        'presupuesto': presupuesto,
        'accion': 'aceptar',
        'titulo': 'Aceptar Presupuesto'
    })

@login_required
def rechazar_presupuesto(request, id):
    """Rechazar un presupuesto (cliente)"""
    presupuesto = get_object_or_404(Presupuesto, id=id)
    
    if presupuesto.cliente != request.user:
        messages.error(request, '❌ Solo el cliente puede rechazar este presupuesto')
        return redirect('lista_presupuestos')
    
    if request.method == 'POST':
        presupuesto.estado = 'rechazado'
        presupuesto.save()
        messages.success(request, '❌ Presupuesto rechazado')
        return redirect('lista_presupuestos')
    
    return render(request, 'gestion/presupuestos/confirmar_accion.html', {
        'presupuesto': presupuesto,
        'accion': 'rechazar',
        'titulo': 'Rechazar Presupuesto'
    })

@login_required
@user_passes_test(puede_gestionar_obras)
def replantear_presupuesto(request, id):
    """Replantear un presupuesto (administrador/constructor)"""
    presupuesto = get_object_or_404(Presupuesto, id=id)
    
    if request.method == 'POST':
        presupuesto.estado = 'replanteado'
        presupuesto.save()
        messages.success(request, '🔄 Presupuesto replanteado')
        return redirect('lista_presupuestos')
    
    return render(request, 'gestion/presupuestos/confirmar_accion.html', {
        'presupuesto': presupuesto,
        'accion': 'replantear',
        'titulo': 'Replantear Presupuesto'
    })

@login_required
def imprimir_presupuesto(request, id):
    """Vista para imprimir presupuesto"""
    presupuesto = get_object_or_404(Presupuesto, id=id)
    items = presupuesto.items.all()
    
    context = {
        'presupuesto': presupuesto,
        'items': items,
    }
    return render(request, 'gestion/presupuestos/imprimir.html', context)

# =============================================
# GESTIÓN DE PERFIL Y USUARIOS
# =============================================

@login_required
def gestionar_perfil(request):
    """Gestión del perfil de usuario"""
    usuario = request.user
    
    if request.method == 'POST':
        # Actualizar información básica
        usuario.first_name = request.POST.get('first_name', '')
        usuario.last_name = request.POST.get('last_name', '')
        usuario.email = request.POST.get('email', '')
        usuario.telefono = request.POST.get('telefono', '')
        usuario.direccion = request.POST.get('direccion', '')
        
        # Manejar avatar
        if request.FILES.get('avatar'):
            usuario.avatar = request.FILES['avatar']
        
        usuario.save()
        
        messages.success(request, '✅ Perfil actualizado exitosamente')
        return redirect('gestionar_perfil')
    
    return render(request, 'gestion/perfil/form.html', {'usuario': usuario})

@login_required
@user_passes_test(es_administrador)
def admin_usuarios(request):
    """Vista personalizada para administrar usuarios - VERSIÓN MEJORADA"""
    usuarios = UsuarioPersonalizado.objects.all().order_by('-date_joined')
    
    # Filtros
    rol_filter = request.GET.get('rol')
    estado_filter = request.GET.get('estado')
    query = request.GET.get('q')
    
    if rol_filter:
        usuarios = usuarios.filter(rol=rol_filter)
    
    if estado_filter:
        usuarios = usuarios.filter(is_active=estado_filter == 'activo')
    
    if query:
        usuarios = usuarios.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )
    
    # Estadísticas para las tarjetas
    total_usuarios = usuarios.count()
    administradores = usuarios.filter(rol='admin')
    constructores = usuarios.filter(rol='constructor')
    clientes = usuarios.filter(rol='cliente')
    vendedores = usuarios.filter(rol='vendedor')
    usuarios_activos = usuarios.filter(is_active=True)
    
    context = {
        'usuarios': usuarios,
        'total_usuarios': total_usuarios,
        'administradores': administradores,
        'constructores': constructores,
        'clientes': clientes,
        'vendedores': vendedores,
        'usuarios_activos': usuarios_activos,
        'roles': UsuarioPersonalizado.ROLES,
        'rol_filter': rol_filter or '',
        'estado_filter': estado_filter or '',
        'query': query or '',
    }
    
    return render(request, 'gestion/usuarios/lista.html', context)

@login_required
@user_passes_test(es_administrador)
def editar_usuario(request, id):
    """Editar usuario (solo admin)"""
    usuario = get_object_or_404(UsuarioPersonalizado, id=id)
    
    if request.method == 'POST':
        usuario.rol = request.POST.get('rol')
        usuario.is_active = request.POST.get('is_active') == 'on'
        usuario.save()
        messages.success(request, f'✅ Usuario {usuario.username} actualizado exitosamente')
        return redirect('admin_usuarios')
    
    return render(request, 'gestion/usuarios/editar.html', {'usuario_editar': usuario})

# =============================================
# VISTAS ADICIONALES Y UTILITARIOS
# =============================================

@login_required
def buscar_ajax(request):
    """Búsqueda AJAX para diferentes modelos"""
    modelo = request.GET.get('modelo')
    query = request.GET.get('q', '')
    
    resultados = []
    if modelo == 'materiales' and query:
        materiales = Material.objects.filter(nombre__icontains=query)[:10]
        resultados = [{'id': m.id, 'nombre': m.nombre, 'precio': str(m.precio)} for m in materiales]
    
    elif modelo == 'obras' and query:
        obras = Obra.objects.filter(nombre__icontains=query)[:10]
        resultados = [{'id': o.id, 'nombre': o.nombre, 'cliente': o.cliente.username} for o in obras]
    
    return JsonResponse({'resultados': resultados})

class RegistroUsuario(CreateView):
    """Registro de nuevos usuarios"""
    form_class = FormularioRegistro
    template_name = 'registration/registro.html'
    success_url = reverse_lazy('home')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        login(self.request, self.object)
        messages.success(self.request, '🎉 ¡Cuenta creada exitosamente!')
        return response

# =============================================
# VISTAS DE REPORTES Y ESTADÍSTICAS
# =============================================

@login_required
@user_passes_test(es_administrador)
def reportes_avanzados(request):
    """Reportes avanzados del sistema"""
    # Estadísticas detalladas
    stats = {
        'total_obras_por_estado': Obra.objects.values('estado').annotate(total=Count('id')),
        'total_presupuestos_por_estado': Presupuesto.objects.values('estado').annotate(total=Count('id')),
        'presupuesto_promedio': Presupuesto.objects.aggregate(avg=Avg('total'))['avg'] or 0,
        'obra_mas_cara': Obra.objects.order_by('-presupuesto_asignado').first(),
    }
    
    return render(request, 'gestion/reportes/avanzados.html', {'stats': stats})
@login_required
def editar_presupuesto(request, id):
    """Editar presupuesto existente"""
    presupuesto = get_object_or_404(Presupuesto, id=id)
    
    # Verificar permisos
    if not (request.user.es_administrador() or presupuesto.cliente == request.user):
        messages.error(request, '❌ No tienes permisos para editar este presupuesto')
        return redirect('lista_presupuestos')
    
    if request.method == 'POST':
        form = FormularioPresupuesto(request.POST, instance=presupuesto)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Presupuesto actualizado exitosamente')
            return redirect('lista_presupuestos')
    else:
        form = FormularioPresupuesto(instance=presupuesto)
    
    return render(request, 'gestion/presupuestos/form.html', {
        'form': form,
        'titulo': 'Editar Presupuesto',
        'accion': 'Actualizar',
        'presupuesto': presupuesto
    })

# =============================================
# DASHBOARD PROFESIONAL PARAGUAY
# =============================================

@login_required
def dashboard_paraguay(request):
    """Dashboard profesional con cálculos en Guaraníes - MEJORADO"""
    from decimal import Decimal
    from django.db.models import Sum, Count, Avg, Q
    from django.utils import timezone
    from datetime import timedelta
    
    print(f"DEBUG: Dashboard cargado para usuario {request.user.username}")
    
    # Fecha actual y rangos
    hoy = timezone.now().date()
    hace_30_dias = hoy - timedelta(days=30)
    hace_7_dias = hoy - timedelta(days=7)
    
    # Estadísticas principales con debug
    try:
        stats = {
            # Totales
            'total_obras': Obra.objects.count(),
            'total_presupuestos': Presupuesto.objects.count(),
            'total_materiales': Material.objects.count(),
            'total_usuarios': UsuarioPersonalizado.objects.count(),
            'total_maquinarias': Maquinaria.objects.count(),
            'total_herramientas': Herramienta.objects.count(),
            
            # Obras por estado
            'obras_planificadas': Obra.objects.filter(estado='planificada').count(),
            'obras_en_proceso': Obra.objects.filter(estado='en_proceso').count(),
            'obras_finalizadas': Obra.objects.filter(estado='finalizada').count(),
            
            # Presupuestos por estado
            'presupuestos_solicitados': Presupuesto.objects.filter(estado='solicitado').count(),
            'presupuestos_aceptados': Presupuesto.objects.filter(estado='aceptado').count(),
            'presupuestos_rechazados': Presupuesto.objects.filter(estado='rechazado').count(),
        }
        print(f"DEBUG: Stats calculadas: {stats}")
    except Exception as e:
        print(f"ERROR calculando stats: {e}")
        stats = {}
        for key in ['total_obras', 'total_presupuestos', 'total_materiales', 'total_usuarios', 'total_maquinarias', 'total_herramientas', 'obras_planificadas', 'obras_en_proceso', 'obras_finalizadas', 'presupuestos_solicitados', 'presupuestos_aceptados', 'presupuestos_rechazados']:
            stats[key] = 0
    
    # Cálculos financieros en Guaraníes
    try:
        # Presupuestos aceptados (ingresos potenciales)
        presupuestos_aceptados = Presupuesto.objects.filter(estado='aceptado')
        total_presupuestos_aceptados = presupuestos_aceptados.aggregate(
            total=Sum('total')
        )['total'] or Decimal('0')
        
        # Obras en proceso (costo real)
        obras_en_proceso = Obra.objects.filter(estado='en_proceso')
        total_costo_obras = obras_en_proceso.aggregate(
            total=Sum('costo_real')
        )['total'] or Decimal('0')
        
        # Materiales en stock (valor inventario)
        materiales_stock = Material.objects.filter(stock__gt=0)
        valor_inventario = sum(
            (material.precio * material.stock) 
            for material in materiales_stock
        )
        
        # Presupuestos del mes
        presupuestos_mes = Presupuesto.objects.filter(
            fecha_creacion__gte=hace_30_dias
        )
        total_presupuestos_mes = presupuestos_mes.aggregate(
            total=Sum('total')
        )['total'] or Decimal('0')
        
        stats.update({
            'total_presupuestos_aceptados': total_presupuestos_aceptados,
            'total_costo_obras': total_costo_obras,
            'valor_inventario': valor_inventario,
            'total_presupuestos_mes': total_presupuestos_mes,
        })
        
    except Exception as e:
        print(f"Error en cálculos financieros: {e}")
        stats.update({
            'total_presupuestos_aceptados': Decimal('0'),
            'total_costo_obras': Decimal('0'),
            'valor_inventario': Decimal('0'),
            'total_presupuestos_mes': Decimal('0'),
        })
    
    # Obras recientes (últimas 5)
    obras_recientes = Obra.objects.all().order_by('-fecha_creacion')[:5]
    
    # Presupuestos recientes (últimos 5)
    presupuestos_recientes = Presupuesto.objects.all().order_by('-fecha_creacion')[:5]
    
    # Actividad reciente (últimos 7 días)
    actividades_recientes = []
    
    # Obras creadas recientemente
    obras_nuevas = Obra.objects.filter(fecha_creacion__gte=hace_7_dias)
    for obra in obras_nuevas:
        actividades_recientes.append({
            'tipo': 'obra',
            'accion': 'creada',
            'objeto': obra.nombre,
            'fecha': obra.fecha_creacion,
            'usuario': obra.creado_por.username,
            'icono': '🏗️'
        })
    
    # Presupuestos creados recientemente
    presupuestos_nuevos = Presupuesto.objects.filter(fecha_creacion__gte=hace_7_dias)
    for presupuesto in presupuestos_nuevos:
        actividades_recientes.append({
            'tipo': 'presupuesto',
            'accion': 'creado',
            'objeto': f"Presupuesto #{presupuesto.id}",
            'fecha': presupuesto.fecha_creacion,
            'usuario': presupuesto.cliente.username,
            'icono': '💰'
        })
    
    # Ordenar actividades por fecha
    actividades_recientes.sort(key=lambda x: x['fecha'], reverse=True)
    actividades_recientes = actividades_recientes[:10]  # Limitar a 10
    
    # Distribución de obras por departamento (Paraguay)
    obras_por_departamento = Obra.objects.exclude(ubicacion__isnull=True).exclude(ubicacion='').values('ubicacion').annotate(
        total=Count('id')
    ).order_by('-total')[:8]
    
    context = {
        'stats': stats,
        'obras_recientes': obras_recientes,
        'presupuestos_recientes': presupuestos_recientes,
        'actividades_recientes': actividades_recientes,
        'obras_por_departamento': list(obras_por_departamento),
        'hoy': hoy,
        'user': request.user,
    }
    
    return render(request, 'gestion/dashboard/dashboard_paraguay.html', context)

@login_required
def crear_presupuesto_avanzado(request):
    """Vista avanzada para crear presupuestos con cálculos en tiempo real"""
    
    # Formset para items
    ItemFormSet = formset_factory(ItemPresupuestoForm, extra=1, can_delete=True)
    
    if request.method == 'POST':
        form = PresupuestoForm(request.POST, user=request.user)
        formset = ItemFormSet(request.POST, prefix='items')
        
        if form.is_valid() and formset.is_valid():
            try:
                # Crear presupuesto
                presupuesto = form.save(commit=False)
                presupuesto.cliente = request.user if request.user.es_cliente else presupuesto.obra.cliente
                presupuesto.estado = 'en_revision'
                presupuesto.save()
                
                # Guardar items
                for item_form in formset:
                    if item_form.cleaned_data and not item_form.cleaned_data.get('DELETE', False):
                        item = item_form.save(commit=False)
                        item.presupuesto = presupuesto
                        item.save()
                
                # Calcular totales automáticamente
                presupuesto.calcular_totales()
                
                messages.success(
                    request, 
                    f'✅ Presupuesto {presupuesto.codigo_presupuesto} creado exitosamente. '
                    f'Total: {presupuesto.total:,.0f} Gs.'
                )
                return redirect('detalle_presupuesto', id=presupuesto.id)
                
            except Exception as e:
                messages.error(request, f'❌ Error al crear el presupuesto: {str(e)}')
    else:
        form = PresupuestoForm(user=request.user)
        formset = ItemFormSet(prefix='items')
    
    context = {
        'form': form,
        'formset': formset,
        'titulo': 'Crear Presupuesto Avanzado',
        'obras': Obra.objects.filter(cliente=request.user) if request.user.es_cliente else Obra.objects.all(),
    }
    
    return render(request, 'gestion/presupuestos/crear_avanzado.html', context)

@login_required
def calcular_presupuesto_ajax(request):
    """Endpoint AJAX para cálculos en tiempo real"""
    if request.method == 'POST' and request.is_ajax():
        try:
            data = json.loads(request.body)
            items = data.get('items', [])
            
            subtotal = Decimal('0')
            items_calculados = []
            
            for item in items:
                cantidad = Decimal(str(item.get('cantidad', 0)))
                precio = Decimal(str(item.get('precio_unitario', 0)))
                total_item = cantidad * precio
                
                subtotal += total_item
                items_calculados.append({
                    'total_item': float(total_item),
                    'total_formateado': f"Gs. {total_item:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
                })
            
            iva_porcentaje = Decimal(str(data.get('iva_porcentaje', 10)))
            iva_monto = (subtotal * iva_porcentaje) / Decimal('100')
            total = subtotal + iva_monto
            
            response_data = {
                'success': True,
                'subtotal': float(subtotal),
                'iva_monto': float(iva_monto),
                'total': float(total),
                'subtotal_formateado': f"Gs. {subtotal:,.0f}".replace(",", "X").replace(".", ",").replace("X", "."),
                'iva_formateado': f"Gs. {iva_monto:,.0f}".replace(",", "X").replace(".", ",").replace("X", "."),
                'total_formateado': f"Gs. {total:,.0f}".replace(",", "X").replace(".", ",").replace("X", "."),
                'items': items_calculados
            }
            
            return JsonResponse(response_data)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
def obtener_datos_material_ajax(request, material_id):
    """Obtener datos de material para autocompletar"""
    try:
        from .models import Material
        material = get_object_or_404(Material, id=material_id)
        
        return JsonResponse({
            'success': True,
            'precio': float(material.precio),
            'unidad_medida': material.unidad_medida,
            'stock': material.stock
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def duplicar_presupuesto(request, presupuesto_id):
    """Duplicar un presupuesto existente"""
    presupuesto_original = get_object_or_404(Presupuesto, id=presupuesto_id)
    
    # Verificar permisos
    if not (request.user.es_administrador or presupuesto_original.cliente == request.user):
        messages.error(request, '❌ No tienes permisos para duplicar este presupuesto.')
        return redirect('lista_presupuestos')
    
    try:
        # Duplicar presupuesto
        presupuesto_nuevo = Presupuesto.objects.create(
            obra=presupuesto_original.obra,
            cliente=presupuesto_original.cliente,
            constructor=presupuesto_original.constructor,
            descripcion_servicios=presupuesto_original.descripcion_servicios,
            iva_porcentaje=presupuesto_original.iva_porcentaje,
            dias_validez=presupuesto_original.dias_validez,
            estado='solicitado'
        )
        
        # Duplicar items
        for item_original in presupuesto_original.items.all():
            ItemPresupuesto.objects.create(
                presupuesto=presupuesto_nuevo,
                tipo=item_original.tipo,
                descripcion=item_original.descripcion,
                cantidad=item_original.cantidad,
                unidad_medida=item_original.unidad_medida,
                precio_unitario=item_original.precio_unitario,
                material=item_original.material,
                maquinaria=item_original.maquinaria,
                herramienta=item_original.herramienta,
                orden=item_original.orden
            )
        
        # Calcular totales
        presupuesto_nuevo.calcular_totales()
        
        messages.success(
            request, 
            f'✅ Presupuesto duplicado exitosamente. Nuevo código: {presupuesto_nuevo.codigo_presupuesto}'
        )
        return redirect('detalle_presupuesto', id=presupuesto_nuevo.id)
        
    except Exception as e:
        messages.error(request, f'❌ Error al duplicar el presupuesto: {str(e)}')
        return redirect('detalle_presupuesto', id=presupuesto_id)

@login_required
def reporte_presupuesto_pdf(request, presupuesto_id):
    """Generar reporte PDF del presupuesto"""
    presupuesto = get_object_or_404(Presupuesto, id=presupuesto_id)
    
    # Verificar permisos
    if not (request.user.es_administrador or presupuesto.cliente == request.user):
        messages.error(request, '❌ No tienes permisos para ver este reporte.')
        return redirect('lista_presupuestos')
    
    context = {
        'presupuesto': presupuesto,
        'items': presupuesto.items.all().order_by('orden', 'tipo'),
        'fecha_actual': timezone.now(),
    }
    
    return render(request, 'gestion/presupuestos/reporte_pdf.html', context)

@login_required
def generar_reporte_presupuesto_pdf(request, presupuesto_id):
    """Generar reporte PDF profesional del presupuesto"""
    presupuesto = get_object_or_404(Presupuesto, id=presupuesto_id)
    
    # Verificar permisos
    if not (request.user.es_administrador or presupuesto.cliente == request.user):
        messages.error(request, '❌ No tienes permisos para ver este reporte.')
        return redirect('lista_presupuestos')
    
    try:
        pdf_generator = PDFGenerator()
        pdf_buffer = pdf_generator.generar_pdf_presupuesto(presupuesto)
        
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="presupuesto_{presupuesto.codigo_presupuesto}.pdf"'
        
        messages.success(request, '✅ Reporte PDF generado exitosamente.')
        return response
        
    except Exception as e:
        messages.error(request, f'❌ Error al generar el PDF: {str(e)}')
        return redirect('detalle_presupuesto', id=presupuesto_id)

@login_required
def generar_reporte_completo_pdf(request, presupuesto_id):
    """Generar reporte completo con gráficos"""
    presupuesto = get_object_or_404(Presupuesto, id=presupuesto_id)
    
    if not (request.user.es_administrador or presupuesto.cliente == request.user):
        messages.error(request, '❌ No tienes permisos para ver este reporte.')
        return redirect('lista_presupuestos')
    
    try:
        # Generar gráficos
        grafico_distribucion = ChartGenerator.generar_grafico_distribucion_presupuesto(presupuesto)
        grafico_evolucion = ChartGenerator.generar_grafico_evolucion_presupuestos(request.user)
        
        context = {
            'presupuesto': presupuesto,
            'items': presupuesto.items.all().order_by('orden', 'id'),
            'grafico_distribucion': grafico_distribucion,
            'grafico_evolucion': grafico_evolucion,
            'fecha_generacion': datetime.now(),
        }
        
        # Renderizar template HTML
        html_string = render_to_string('gestion/presupuestos/reporte_completo_pdf.html', context)
        
        # Generar PDF
        pdf_file = HTML(string=html_string).write_pdf()
        
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_completo_{presupuesto.codigo_presupuesto}.pdf"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'❌ Error al generar el reporte completo: {str(e)}')
        return redirect('detalle_presupuesto', id=presupuesto_id)

@login_required
def reportes_estadisticas_avanzadas(request):
    """Vista de reportes y estadísticas avanzadas"""
    from django.db.models import Count, Sum, Avg, Q
    from django.utils import timezone
    
    # Filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_reporte = request.GET.get('tipo_reporte', 'general')
    
    # Fechas por defecto (últimos 30 días)
    if not fecha_inicio:
        fecha_inicio = (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not fecha_fin:
        fecha_fin = timezone.now().strftime('%Y-%m-%d')
    
    # Construir filtros
    filtros = Q(fecha_creacion__date__gte=fecha_inicio, fecha_creacion__date__lte=fecha_fin)
    
    if not request.user.es_administrador:
        filtros &= Q(cliente=request.user)
    
    # Estadísticas generales
    presupuestos = Presupuesto.objects.filter(filtros)
    obras = Obra.objects.filter(filtros) if request.user.es_administrador else Obra.objects.filter(cliente=request.user)
    
    stats = {
        'total_presupuestos': presupuestos.count(),
        'presupuestos_aceptados': presupuestos.filter(estado='aceptado').count(),
        'presupuestos_rechazados': presupuestos.filter(estado='rechazado').count(),
        'total_obras': obras.count(),
        'obras_en_proceso': obras.filter(estado='en_proceso').count(),
        'monto_total_presupuestado': presupuestos.aggregate(total=Sum('total'))['total'] or Decimal('0'),
        'monto_total_aceptado': presupuestos.filter(estado='aceptado').aggregate(total=Sum('total'))['total'] or Decimal('0'),
    }
    
    # Datos para gráficos
    presupuestos_por_estado = presupuestos.values('estado').annotate(
        cantidad=Count('id'),
        monto_total=Sum('total')
    ).order_by('-monto_total')
    
    presupuestos_por_mes = presupuestos.extra({
        'mes': "strftime('%%Y-%%m', fecha_creacion)"
    }).values('mes').annotate(
        cantidad=Count('id'),
        monto=Sum('total')
    ).order_by('mes')
    
    # Top materiales más usados
    materiales_populares = Material.objects.annotate(
        veces_usado=Count('itempresupuesto')
    ).filter(veces_usado__gt=0).order_by('-veces_usado')[:10]
    
    context = {
        'stats': stats,
        'presupuestos_por_estado': list(presupuestos_por_estado),
        'presupuestos_por_mes': list(presupuestos_por_mes),
        'materiales_populares': materiales_populares,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'tipo_reporte': tipo_reporte,
        'user': request.user,
    }
    
    return render(request, 'gestion/reportes/estadisticas_avanzadas.html', context)

@login_required
def exportar_estadisticas_excel(request):
    """Exportar estadísticas completas a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime
    
    # Crear workbook
    wb = Workbook()
    
    # HOJA 1: OBRAS
    ws_obras = wb.active
    ws_obras.title = 'Obras'
    
    # Encabezados
    headers_obras = ['Nombre', 'Ubicación', 'Estado', 'Cliente', 'Presupuesto Asignado', 'Costo Real', 'Fecha Inicio', 'Progreso %']
    ws_obras.append(headers_obras)
    
    # Estilo encabezados
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for cell in ws_obras[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos obras
    obras = Obra.objects.all().order_by('-fecha_creacion')
    for obra in obras:
        ws_obras.append([
            obra.nombre,
            obra.ubicacion or 'Sin ubicación',
            obra.get_estado_display(),
            obra.cliente.get_full_name() or obra.cliente.username,
            float(obra.presupuesto_asignado),
            float(obra.costo_real),
            obra.fecha_inicio.strftime('%d/%m/%Y'),
            obra.progreso_porcentaje()
        ])
    
    # Ajustar anchos
    for i, col in enumerate(ws_obras.columns, 1):
        ws_obras.column_dimensions[get_column_letter(i)].width = 20
    
    # HOJA 2: PRESUPUESTOS
    ws_presup = wb.create_sheet('Presupuestos')
    headers_presup = ['Código', 'Obra', 'Cliente', 'Estado', 'Subtotal', 'IVA', 'Total', 'Fecha Creación']
    ws_presup.append(headers_presup)
    
    for cell in ws_presup[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    presupuestos = Presupuesto.objects.all().order_by('-fecha_creacion')
    for presup in presupuestos:
        ws_presup.append([
            presup.codigo_presupuesto,
            presup.obra.nombre,
            presup.cliente.get_full_name() or presup.cliente.username,
            presup.get_estado_display(),
            float(presup.subtotal),
            float(presup.iva_monto),
            float(presup.total),
            presup.fecha_creacion.strftime('%d/%m/%Y %H:%M')
        ])
    
    for i, col in enumerate(ws_presup.columns, 1):
        ws_presup.column_dimensions[get_column_letter(i)].width = 20
    
    # HOJA 3: MATERIALES
    ws_mat = wb.create_sheet('Materiales')
    headers_mat = ['Nombre', 'Descripción', 'Precio', 'Stock', 'Unidad', 'Valor Total']
    ws_mat.append(headers_mat)
    
    for cell in ws_mat[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    materiales = Material.objects.all().order_by('nombre')
    for mat in materiales:
        ws_mat.append([
            mat.nombre,
            mat.descripcion[:50],
            float(mat.precio),
            mat.stock,
            mat.unidad_medida,
            float(mat.precio * mat.stock)
        ])
    
    for i, col in enumerate(ws_mat.columns, 1):
        ws_mat.column_dimensions[get_column_letter(i)].width = 20
    
    # HOJA 4: ESTADÍSTICAS
    ws_stats = wb.create_sheet('Estadísticas')
    ws_stats.append(['ESTADÍSTICAS GENERALES DEL SISTEMA'])
    ws_stats['A1'].font = Font(bold=True, size=16, color='667eea')
    ws_stats.append([])
    
    stats_data = [
        ['Total Obras', Obra.objects.count()],
        ['Obras Planificadas', Obra.objects.filter(estado='planificada').count()],
        ['Obras En Proceso', Obra.objects.filter(estado='en_proceso').count()],
        ['Obras Finalizadas', Obra.objects.filter(estado='finalizada').count()],
        [],
        ['Total Presupuestos', Presupuesto.objects.count()],
        ['Presupuestos Solicitados', Presupuesto.objects.filter(estado='solicitado').count()],
        ['Presupuestos Aceptados', Presupuesto.objects.filter(estado='aceptado').count()],
        ['Presupuestos Rechazados', Presupuesto.objects.filter(estado='rechazado').count()],
        [],
        ['Total Materiales', Material.objects.count()],
        ['Total Usuarios', UsuarioPersonalizado.objects.count()],
    ]
    
    for row in stats_data:
        ws_stats.append(row)
    
    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 20
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_completo_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def exportar_reporte_completo(request):
    """Vista para exportar reporte completo - alias de exportar_estadisticas_excel"""
    return exportar_estadisticas_excel(request)

def chat_widget(request):
    """Vista para el widget de chat flotante"""
    return render(request, 'chat/chat_widget.html')

def offline(request):
    """Página offline para PWA"""
    return render(request, 'offline.html')

# =============================================
# VISTAS PWA (Progressive Web App)
# =============================================

def service_worker(request):
    """Service Worker para PWA"""
    import os
    from django.conf import settings
    
    sw_path = os.path.join(settings.BASE_DIR, 'static', 'js', 'sw.js')
    try:
        with open(sw_path, 'r', encoding='utf-8') as f:
            content = f.read()
        response = HttpResponse(content, content_type='application/javascript')
        response['Cache-Control'] = 'no-cache'
        return response
    except FileNotFoundError:
        return HttpResponse('// Service Worker no encontrado', content_type='application/javascript')

def manifest_json(request):
    """Manifest JSON para PWA"""
    import json
    from pwa_config import PWA_MANIFEST
    
    manifest = PWA_MANIFEST.copy()
    
    response = HttpResponse(json.dumps(manifest, indent=2), content_type='application/json')
    response['Cache-Control'] = 'no-cache'
    return response

def pwa_install(request):
    """Página de instalación PWA"""
    return render(request, 'pwa_install.html')